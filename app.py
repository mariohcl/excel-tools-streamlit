import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Herramientas Excel", layout="centered")
st.title("🔧 Limpieza de BBDD")

opcion = st.radio("Selecciona una funcionalidad:", [
    "📊 Comparar dos archivos Excel por email",
    "🧹 Limpiar duplicados en un Excel"
])

if opcion == "📊 Comparar dos archivos Excel por email":
    st.subheader("📁 Subir los archivos")
    archivo1 = st.file_uploader("Archivo 1 a comparar (Debe tener la columna llamada 'Email')", type=["xlsx"], key="file1")
    archivo2 = st.file_uploader("Archivo 2 nuestra base (Debe tener la columna llamada 'E-mail')", type=["xlsx"], key="file2")

    if archivo1 and archivo2:
        df1 = pd.read_excel(archivo1)
        df2 = pd.read_excel(archivo2)

        df1['Email'] = df1['Email'].astype(str).str.strip().str.lower()
        df2['E-mail'] = df2['E-mail'].astype(str).str.strip().str.lower()
        df1['Estatus'] = df1['Email'].isin(df2['E-mail']).map({True: 'SI', False: 'NO'})

        salida = "comparacion_resultado.xlsx"
        df1.to_excel(salida, index=False)

        wb = load_workbook(salida)
        ws = wb.active
        col_idx = [c.value for c in ws[1]].index("Estatus") + 1
        verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        for row in ws.iter_rows(min_row=2):
            if row[col_idx - 1].value == "SI":
                for c in row:
                    c.fill = verde

        wb.save(salida)

        with open(salida, "rb") as f:
            st.download_button("📥 Descargar archivo con estatus", f, file_name=salida)

elif opcion == "🧹 Limpiar duplicados en un Excel":
    st.subheader("📁 Subir archivo con emails repetidos")
    archivo = st.file_uploader("Archivo Excel (columna 'E-mail')", type=["xlsx"], key="file3")

    if archivo:
        df = pd.read_excel(archivo)
        df['E-mail'] = df['E-mail'].astype(str).str.strip().str.lower()
        df = df[df['E-mail'].notna() & (df['E-mail'] != '')]

        conteo = df['E-mail'].value_counts().reset_index()
        conteo.columns = ['E-mail', 'Repeticiones']

        df_limpio = df.drop_duplicates(subset='E-mail', keep='first')
        df_final = pd.merge(df_limpio, conteo, on='E-mail', how='left')

        cols = df_final.columns.tolist()
        cols.remove('Repeticiones')
        df_final = df_final[['Repeticiones'] + cols]

        salida = "limpieza_resultado.xlsx"
        df_final.to_excel(salida, index=False)

        wb = load_workbook(salida)
        ws = wb.active
        rojo = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for row in ws.iter_rows(min_row=2):
            if row[0].value > 1:
                for c in row:
                    c.fill = rojo

        wb.save(salida)

        with open(salida, "rb") as f:
            st.download_button("📥 Descargar archivo limpio", f, file_name=salida)
